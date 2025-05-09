# -*- coding: utf-8 -*-
from PyQt6.QtWidgets import QMainWindow,  QApplication
from PyQt6.QtCore import pyqtSlot
import sys, math, efforts
import tkinter.messagebox
import numpy as np
from Ui_main import Ui_MainWindow
from fig_press import Form as Form2
from fig_suck import Form as Form1

class MainWindow(QMainWindow, Ui_MainWindow):
    #主窗口
    #以列作为维度的二维矩阵的数据点结果
    points_suck=np.array([])
    points_press=np.array([])
    points_back=np.array([])
    points_front=np.array([])
    points=np.array([])#生成点
    points0=np.array([])#导入点
    sw0=np.array([])#导入点通道
    loc='./database.xlsx'
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)
    
        
        self.datas=efforts.load(self.loc)
        self.spinBox.setMaximum(len(self.datas))

        n=self.spinBox.value() 
        
        self.model=self.datas[n]  #模式导入
        self.textBrowser.setText(str(self.model))
        self.textBrowser_2.setText(str(self.model['备注']))
        self.textBrowser_3.setText(str(self.model['入口宽度']))
        
        
        
        #link的信号
        self.suctionsurface.clicked.connect(self.open_suck)
        self.pressuresurface.clicked.connect(self.open_press)
        self.leadingedge.clicked.connect(self.open_front)
        self.trailingedge.clicked.connect(self.open_back)
        self.pushButton_5.clicked.connect(self.open_pass)
        
        
        self.on_pushButton_2_clicked()
    
    def open_pass(self):
        p0=self.points0#获得导入点的通道
        p0x=p0[:, 0].tolist()
        p0y=p0[:, 1].tolist()
        a=p0x.index(min(p0x))
        b=p0x.index(max(p0x))
        p=float(self.textBrowser_3.toPlainText())
        #切割曲线（逆时针则为函数曲线）#
        l1=[]#上半
        l2=[]#下半
        if a<b:
            l2=np.array([p0x[a:b+1], p0y[a:b+1]]).T.tolist()
            l1=np.array([p0x[b:]+p0x[0:a+1], p0y[b:]+p0y[0:a+1]]).T.tolist()
        else:
            l1=np.array([p0x[b:a+1], p0y[b:a+1]]).T.tolist()
            l2=np.array([p0x[a:]+p0x[0:b+1], p0y[a:]+p0y[0:b+1]]).T.tolist()
        l1.reverse()
        l20=np.array(l2)
        l20[:, 1]=l20[:, 1]+p
        l2=l20.tolist()

        #计算通道
        s=[0]#弧位置
        w=[p]#通道宽度
        centerpass=[]#中弧线位置
        for i in range(1, len(l1)-1):
            pl1=[l1[i][0], l1[i][1]]#选定压力面的某点
            s=s+[s[i-1]+math.hypot(l1[i][0]-l1[i-1][0],l1[i][1]-l1[i-1][1])]
            dl1=(l1[i+1][1]-l1[i-1][1])/(l1[i+1][0]-l1[i-1][0])#该点斜率
            thita1=math.atan(dl1)
            dthita=[]
            for j in range(1, len(l2)-1):
                #遍历吸力面的所有点
                dp=[l2[j][0]-l1[i][0], l2[j][1]-l1[i][1]]#两点相对位置
                if dp[0]!=0:
                    thita0=math.atan(dp[1]/dp[0])
                else:
                    thita0=math.pi/2
                thita2=math.atan((l2[j+1][1]-l2[j-1][1])/(l2[j+1][0]-l2[j-1][0]))
                dthita1=math.pi-(thita0-thita1)
                dthita2=thita0-thita2
                dthita=dthita+[abs(dthita2-dthita1)]
            loc=dthita.index(min(dthita))
            pl2=[l2[loc][0], l2[loc][1]]
            w=w+[math.hypot(pl2[1]-pl1[1],pl2[0]-pl1[0])]
            centerpass=centerpass+[[(pl2[0]+pl1[0])/2,(pl2[1]+pl1[1])/2]]
        s=s+[s[-1]+s[-1]-s[-2]]
        w=w+[p]
        self.fig_main.mpl.plots(centerpass)
        self.textBrowser_4.setText(str(min(w)))
        self.sw0=np.array([s, w]).T
        #绘图-通道
        from pass0 import Form
        form=Form(self)
        form.sw0=self.sw0
        cur1=efforts.curve(np.array(l1))
        cur2=efforts.curve(np.array(l2))
        form.sw0_suc=cur1
        form.sw0_pre=cur2
        form.on_pushButton_clicked()
        form.show()
        
    def open_back(self):
        from controler_back import Form
        form=Form(self)
        md=self.model
        form.doubleSpinBox_x.setValue(float(md['后缘端点横坐标']))
        form.doubleSpinBox_y.setValue(float(md['后缘端点纵坐标']))
        form.circle_x.setValue(float(md['后缘圆心横坐标']))
        form.circle_y.setValue(float(md['后缘圆心纵坐标']))
        form.circle_x_2.setValue(float(md['后缘圆弧度数']))
        
        form.set()
        form.Signal_model.connect(self.emit_back)
        form.Signal_model2.connect(self.draw_back)
        form.show()
        
    def emit_back(self, p):
        #导出
        self.model.update(p)
    def draw_back(self, p):
        #绘制
        m0=self.model.copy()
        self.model.update(p)
        self.on_pushButton_2_clicked()
        self.model=m0

    
    def open_front(self):
        from controler_front import Form
        form=Form(self)
        md=self.model
        form.doubleSpinBox_x.setValue(float(md['前缘端点横坐标']))
        form.doubleSpinBox_y.setValue(float(md['前缘端点纵坐标']))
        form.circle_x.setValue(float(md['前缘圆心横坐标']))
        form.circle_y.setValue(float(md['前缘圆心纵坐标']))
        form.circle_x_2.setValue(float(md['前缘圆弧度数']))
        
        form.set()
        form.Signal_model.connect(self.emit_front)
        form.Signal_model2.connect(self.draw_front)
        form.show()
    def emit_front(self, p):
        #导出
        self.model.update(p)
    def draw_front(self, p):
        #绘制
        m0=self.model.copy()
        self.model.update(p)
        self.on_pushButton_2_clicked()
        self.model=m0


        
        

    def open_suck(self):
        #打开吸力面绘图界面
        form = Form1(self)
        form.model=self.model
        form.on_pushButton_3_clicked()
        form.points0=self.points0
        
        r1=self.points_front[0, :]
        r2=self.points_back[-1, :]
        
        md=self.model
        fronpoint_suck=np.array([float(md['前缘端点横坐标']), float(md['前缘端点纵坐标'])])
        backpoint_suck=np.array([float(md['后缘端点横坐标']), float(md['后缘端点纵坐标'])])
        
        a1=r1-fronpoint_suck
        form.k1=a1[1]/a1[0]
        a2=r2-backpoint_suck
        form.k2=a2[1]/a2[0]  #校正参数
        
        form.point1x=str(r1[0])
        form.point1y=str(r1[1])
        form.point2x=str(r2[0])
        form.point2y=str(r2[1])
        
        form.Signal_draw.connect(self.fig_suck_points)#绘图
        form.Signal_model.connect(self.fig_suck_model)#改变模式
        form.show()
        
    def fig_suck_model(self, dic):
        self.model.update(dic)
    def fig_suck_points(self, p1):
        #在主窗口里绘图
        self.fig_main.mpl.plots(p1)
        #p是二维列表
        


    def open_press(self):
        #打开压力面绘图界面
        form = Form2(self)
        form.model=self.model
        form.on_pushButton_3_clicked()
        form.points0=self.points0
        
        r1=self.points_front[-1, :]
        r2=self.points_back[0, :]
        
        md=self.model
        fronpoint_suck=np.array([float(md['前缘端点横坐标']), float(md['前缘端点纵坐标'])])
        backpoint_suck=np.array([float(md['后缘端点横坐标']), float(md['后缘端点纵坐标'])])
        
        a1=r1-fronpoint_suck
        form.k1=a1[1]/a1[0]
        a2=r2-backpoint_suck
        form.k2=a2[1]/a2[0]  #校正参数
        
        
        form.point1x=str(r1[0])
        form.point1y=str(r1[1])
        form.point2x=str(r2[0])
        form.point2y=str(r2[1])
        
        form.Signal_draw.connect(self.fig_press_points)
        form.Signal_model.connect(self.fig_press_model)#改变模式
        form.show()
    def fig_press_model(self, dict):
        self.model.update(dict)


    def fig_press_points(self, p):
        #在主窗口里绘图
        self.fig_main.mpl.plots(p)
  
        
    #主窗口功能
    def toarray(self, points):
        if type(points)==np.ndarray:
            pass
        elif type(points)==list:
            points=np.array(points)
        return points  #两列二维数组
    #画圆弧
    def circle(self, thita, p0, r0, points:int):
            #输入：圆弧度数，端点位置，圆心位置，圆弧段点数
            thita=math.radians(thita)
            dr=self.toarray(r0)-self.toarray(p0)
            thita0=math.atan2(dr[1], dr[0])
            th=[thita0-thita/2+math.pi,thita0+thita/2+math.pi]
            angle=np.linspace(min(th), max(th), points)
            dx=list(map(math.cos, angle.tolist()))
            dy=list(map(math.sin, angle.tolist()))
            dthita=np.array([dx, dy]).T
            #半径
            r=math.hypot(dr[0], dr[1])*math.cos(thita/2)
            
            finalpoints=np.array([r0]*points)+r*dthita
            self.fig_main.mpl.scatters(r0)
            self.fig_main.mpl.scatters(p0)
            return finalpoints
                
    @pyqtSlot()
    #绘制
    def on_pushButton_2_clicked(self):
        #绘图  前后缘  坐标变换
        #self.textBrowser_2.setText(self.model['备注'])
        #self.textBrowser_3.setText(self.model['入口宽度'])
        md=self.model
        
        #前缘
        self.points_front=self.circle(float(md['前缘圆弧度数']),\
        np.array([float(md['前缘端点横坐标']), float(md['前缘端点纵坐标'])]), \
        np.array([float(md['前缘圆心横坐标']), float(md['前缘圆心纵坐标'])]), 40)   
        #后缘
        self.points_back=self.circle(float(md['后缘圆弧度数']),\
        np.array([float(md['后缘端点横坐标']), float(md['后缘端点纵坐标'])]), \
        np.array([float(md['后缘圆心横坐标']), float(md['后缘圆心纵坐标'])]), 40)
        
        #吸力面控制点
        suc_ax=[self.points_front[0, 0]]+efforts.choose(md['吸力面控制点横坐标'])+\
        [self.points_back[-1, 0]]
        
        suc_ay=[self.points_front[0, 1]]+efforts.choose(md['吸力面控制点纵坐标'])+\
        [self.points_back[-1, 1]]
        suc_wn=efforts.choose(md['吸力面控制点权重'])
        suc_kn=efforts.choose(md['吸力面节点矢量'])
        suc_degree=int(efforts.choose(md['吸力面阶数'])[0])
        
        #压力面控制点
        pre_ax=[self.points_front[-1, 0]]+efforts.choose(md['压力面控制点横坐标'])+\
        [self.points_back[0, 0]]
        
        pre_ay=[self.points_front[-1, 1]]+efforts.choose(md['压力面控制点纵坐标'])+\
        [self.points_back[0, 1]]
        pre_wn=efforts.choose(md['压力面控制点权重'])
        pre_kn=efforts.choose(md['压力面节点矢量'])
        pre_degree=int(efforts.choose(md['压力面阶数'])[0])
        
        
        #吸力面与压力面
        self.points_suck=self.fig_main.mpl.nurbs_plot(suc_ax, suc_ay, suc_wn, suc_kn, suc_degree)
        self.points_press=self.fig_main.mpl.nurbs_plot(pre_ax, pre_ay, pre_wn, pre_kn, pre_degree)
        


        self.fig_main.mpl.plots(self.points_front)
        self.fig_main.mpl.plots(self.points_back)
        if self.points0.size>0:
            self.fig_main.mpl.plots(self.points0)
            p=self.points0.copy()
            p[:, 1]=p[:, 1]+float(self.textBrowser_3.toPlainText())
            self.fig_main.mpl.plots(p)
            
        m=self.points_suck.tolist()
        m.reverse()
        p=self.points_press.tolist()+self.points_back.tolist()+m+self.points_front.tolist()
        self.points=np.array(p)    
        
        #监视
        self.textBrowser.setText(str(md))
        self.fig_main.mpl.draw()

    @pyqtSlot()
    def on_pushButton_3_clicked(self):
        #清除
        self.fig_main.mpl.clear()
        self.fig_main.mpl.draw()
        

    @pyqtSlot()
    def on_actionsave_points_of_results_triggered(self):
        # 导出曲线坐标点
        
        p=self.points.tolist()
        #排序算法暂未编写

        with open('曲线坐标点.txt', 'w') as f:
            f.write(str(len(p)))
            f.write('\n')
            for i in range(len(p)):
                f.write(str(p[i][0]))
                f.write('\t')
                f.write(str(p[i][1]))
                f.write('\t')
                f.write(str(0))
                f.write('\n')
        msg='''导出曲线样本点成功'''
        
        tkinter.messagebox.showinfo(title='你真牛逼', message=msg) 
        
        
        
    @pyqtSlot()
    def on_actionsave_the_modle_triggered(self):
        # 导出样式
        from openpyxl import load_workbook
        wb = load_workbook(self.loc)
        ws = wb.worksheets[0]
        maxrow=ws.max_row  
        maxcol=ws.max_column  
        
        self.model['备注']=self.textBrowser_2.toPlainText()
        self.model['入口宽度']=self.textBrowser_3.toPlainText()
        md=self.model
        for i in range(1, maxrow+1):
            if i==1:
                ws.cell(i, maxcol+1).value=str(maxcol)
            else:
                ws.cell(i, maxcol+1).value=str(md[ws.cell(i, 1).value])
                
    
        wb.save('database.xlsx')
        

        
        #重置
        self.datas=efforts.load(self.loc)
        self.spinBox.setMaximum(len(self.datas))

        msg='''导出样式成功'''
        tkinter.messagebox.showinfo(title='原神，启动！', message=msg) 

    @pyqtSlot(int)
    def on_spinBox_valueChanged(self, p0):
        self.model=self.datas[p0]
   

    @pyqtSlot()
    def on_action11_triggered(self):
        #导入坐标点
        import tkinter as tk
        from tkinter import filedialog
        path = filedialog.askopenfilename()
        path0=path.split('.')
        m1=[]
        if path0[-1]=='txt':
            with open(path, 'r') as f:
            #请选择固定格式的txt文件
                m=f.readlines()
                n=0
                for i in m:
                    try:
                        #
                        mk=efforts.choose(i)[0:2]
                        if len(mk)!=1:
                            m1=m1+[mk]
                    except:
                        n=n+1
                        if n>=4:
                            tk.messagebox.showerror(title='WARNING', message='请输入规定格式的数据') 
                            break
           
                        
        elif path0[-1]=='xlsx'or'xls':
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws=wb.active
            p=ws.rows
            for i in p:
                if len(i)!=1:
                    try:
                        m1=m1+[[float(i[0].value), float(i[1].value)]]
                    except:
                        tk.messagebox.showerror(title='WARNING', message='请输入规定格式的数据') 
                        break
        elif path0=='':
            pass
        else:
            tk.messagebox.showerror(title='WARNING', message='目前仅支持txt,xls,xlsx格式') 

        try:
            m2=np.array(m1)
            self.fig_main.mpl.plots(m2)
            self.points0=m2
            tk.messagebox.showinfo(title='6', message='导入成功')
        except:
            tk.messagebox.showerror(title='g', message='导入失败')
            pass

    @pyqtSlot()
    def on_action33_triggered(self):
        msg='''欢迎使用'''
        tkinter.messagebox.showinfo(title='你真牛逼', message=msg) 


            
            
            



if __name__ == "__main__":
    app = QApplication(sys.argv)
    ui = MainWindow()
    ui.show()
    sys.exit(app.exec())
        
