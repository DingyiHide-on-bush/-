import numpy as np
from geomdl import NURBS

def yuan(rou,thita,n):
    #过原点，圆心极坐标为（rou，thita），存在于x轴上方的圆弧段
    thita0 = np.linspace(-thita, np.pi+thita,n)
    x_cos = np.cos(thita0)*rou
    y_sin = np.sin(thita0)*rou
    x=x_cos+rou*float(np.cos(thita))
    y=y_sin+rou*float(np.sin(thita))

    line_yuan=np.array([x,y]).T
    return line_yuan

def turn_around(thita):
    #生成旋转矩阵(2*2)
    A=np.array([[np.cos(thita),np.sin(thita)],[-np.sin(thita),np.cos(thita)]])
    return A


def nurbs(an:list,weigh:list, knot:list, degree:int):
    
    curve = NURBS.Curve()
    curve.degree = degree  #阶数一定是整数，最小控制点数=阶数+1
    # Set control points (weights vector will be 1 by default)
    # Use curve.ctrlptsw is if you are using homogeneous points as Pw
    curve.ctrlpts = an
    curve.weights=weigh
    #节点矢量元素个数=阶数+控制点数+1，其内元素从0不严格单增至1
    #为了让曲线经过首尾控制点，首末节点均需具备（阶数+1）个重复度
    curve.knotvector = [0]*(degree+1)+knot+[1]*(degree+1)
    #curve.knotvector = [0, 1/7, 2/7, 3/7, 4/7, 5/7, 6/7, 1]#元素间隔代表了其附近的均匀性，
    # Set evaluation delta (controls the number of curve points)
    curve.delta = 0.005
    # Get curve points (the curve will be automatically evaluated)
    curve_points = np.array(curve.evalpts)
    return curve_points
    
def choose( s0:str):
    #数据处理
    s1=[]
    list0=['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '-', '.']
    for i in s0:
        if i in list0:
            s1=s1+list(i)
        else:
            s1=s1+list('*')
    s1=''.join(s1)
    s=list(filter(None, s1.split('*')))
    s=list(map(float, s))
    return s
    
def load(loc):
    #读取，不修改 
    from openpyxl import load_workbook
    wb = load_workbook(loc)
    ws = wb.worksheets[0]
    # 获得最大列和最大行
    maxrow=ws.max_row
    maxcol=ws.max_column
    database={}
    for j in range(1, maxcol):
        mark=[]
        value=[]
        model={}
        for i in range(1, maxrow+1):
            mark=str(ws.cell(i, 1).value)
            value=str(ws.cell(i, j+1).value)
            model[mark]=value
        database[j]=model
    return database
    

def curve(p):
    import math
    cur=[]
    s=0
    for i in range(2, len(p)):
        a1=math.atan((p[i-1, 1]-p[i-2, 1])/(p[i-1, 0]-p[i-2, 0]))
        a2=math.atan((p[i-1, 1]-p[i, 1])/(p[i-1, 0]-p[i, 0]))
        s1=math.hypot((p[i-1, :]-p[i-2, :])[0], (p[i-1, :]-p[i-2, :])[1])
        if a1!=a2:
            c=abs(s1/(a1-a2))
            s=s+s1
            cur=cur+[[s, c]]
    return cur
    

if __name__ == "__main__":
    load(1)
    
